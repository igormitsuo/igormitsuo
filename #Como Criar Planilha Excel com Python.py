#Relatorio diario- Excel com Python

from openpyxl import Workbook
from time import sleep
from datetime import datetime

#aba=arquivo.active
#aba.title="MODELO"

def menu():
 print('''
********** RELATORIO DIARIO DE ATIVIDADES **********

****** ESCOLHA A OPERAÇÃO DESEJADA ******
[0] SAIR
[1] CRIAR NOVA ABA
[2] PRENCHER RELATORIO
[3] EDITAR RELATORIO
[4] EXCLUIR RELATORIO
[5] EXCLUIR ABA
[6] VER ABAS CRIADAS
       
**************************************
''')
 
 valor=int(input("Entre com a opção \n"))
 if valor not in [0,1,2,3,4,5,6]:
    print("## OPÇÃO INVALIDA: DIGITE UM VALIDO##")
    sleep(1)
 return (valor)

def criar_nova_aba():
 for aba in arquivo:
  print(aba)
 nome_aba=str(input("Digite o nome da nova aba\n"))
 arquivo.create_sheet(nome_aba)
 arquivo.active
 arquivo.save("relatorio_diario.xlsx")
 

def preencher_relatorio():
 pass
def  editar_relatorio():
 pass
def  excluir_relatorio():
 pass
def  excluir_aba():
 pass
def ver_abas_criadas():
 pass




while True:
 opção=menu()
 if opção==0:
  print("OPÇÃO SAIR")
  sleep(2)
  break
 elif opção==1:
  criar_nova_aba()
 elif opção==2:
  preencher_relatorio()
 elif opção==3:
  editar_relatorio()
 elif opção==4:
  excluir_relatorio()
 elif opção==5:
  excluir_aba()
 elif opção==6:
  ver_abas_criadas()
arquivo.save("relatorio_diario.xlsx")

'''data="hoje"
arquivo=Workbook()
aba=arquivo.active
aba.title=f'{data}'
arquivo.create_sheet("amanha")
arquivo.create_sheet("depois")
arquivo.create_sheet("depois_amanha")'''

    
    

#arquivo=Workbook()


#arquivo.save("relatorio_diario.xlsx")

